import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import os
import numpy as np
import re


def number_to_string(num):
    return str(num).zfill(2)

def list_files_in_directory(directory):
    """Lists files."""
    try:
        # Get the list of files in the directory
        files = [file for file in os.listdir(directory) if os.path.isfile(os.path.join(directory, file))]

        # Sort the files alphabetically
        files_sorted = sorted(files)
        
    except FileNotFoundError:
        print(f"The directory {directory} does not exist.")
    except PermissionError:
        print(f"Permission denied to access the directory {directory}.")
    except Exception as e:
        print(f"An error occurred: {e}")
    return files_sorted

def read_eflux_file(file:str):
    # Read the file into a DataFrame, starting from line 2 and ending at line 7681, skipping the first line
    # Since we are skipping the first line (index 0), we start from index 1 (second line) and read 7680 lines
    df_eflux = pd.read_csv(file, header=None, skiprows=1, nrows=7680, delim_whitespace=True, on_bad_lines='skip')

    # Define new headers
    df_eflux.columns = ['MLT', 'ML', '[mW m^-2]']

    return df_eflux

def read_IMF_file():
    #read IMF data
    file_path_IMF = r"G:\My Drive\Butler\Research\NASA project\GitHub\AuroraNSL_2024\AuroraNSL_2024\data\St patrick's day 2015\20150317_IMF.txt"

    # Read the file into a DataFrame without headers
    df_IMF = pd.read_csv(file_path_IMF, header=None, delim_whitespace=True)

    # Define new headers
    headers = ["Year", "Month", "Day", "Hour", "Min", "Sec", "Msec", "Bx[nT]", "By[nT]", "Bz[nT]", "Vx[km/s]", "Vy[km/s]", "Vz[km/s]", "N[cm^(-3)]", "T[Kelvin]"]

    # Set the column headers
    df_IMF.columns = headers

    return df_IMF

def df_to_openpyxl(dataframe,workbook):
    # Convert the dataframe to an openpyxl object
    rows = dataframe_to_rows(dataframe, index=False, header=True)

    # Create a new workbook and worksheet
    worksheet = workbook.active
    
    # Add the rows to the worksheet
    for rnum, row in enumerate(rows, start=0):
        for col_num , the_cell in enumerate(row, start=0):
            worksheet.cell(row=rnum+1, column=col_num+1, value=the_cell)
    
    return worksheet

def namedtuples_to_array(namedtuples):
    
    namedtuple_dict = namedtuples._asdict()
    namedtuple_dict_values = list(namedtuple_dict.values())
    namedtuple_arr = np.array(namedtuple_dict_values)
       
    return namedtuple_arr

def extract_time(s):
    # Use a regular expression to find a pattern where digits precede "UT"
    match = re.search(r'(\d+)UT', s)
    if match:
        time_code = match.group(1)
        
        # Slice to get hour and minute
        hour = int(time_code[:2])
        minute = int(time_code[2:])
        
        return hour, minute

    # Return None if no match is found (optional, depends on how you want to handle errors)
    return None

def extract_datetime_info(string:str):
    # Regex pattern to extract year, month, day, hour, and minute
    pattern = r'(\d{4})(\d{2})(\d{2})_(\d{2})(\d{2})UT'
    match = re.search(pattern, string)
    
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        day = int(match.group(3))
        hour = int(match.group(4))
        minute = int(match.group(5))
        
        return {
            'year': year,
            'month': month,
            'day': day,
            'hour': hour,
            'minute': minute
        }

    return None

year = '2015'
month = '03'
day = '17'
hour = 0
minute = 0

#start to read eFlux files
file_dir_eFlux = r"G:\My Drive\Butler\Research\NASA project\GitHub\AuroraNSL_2024\AuroraNSL_2024\data\St patrick's day 2015\Eflux"
file_dir_main = r"G:\My Drive\Butler\Research\NASA project\GitHub\AuroraNSL_2024\AuroraNSL_2024\data\St patrick's day 2015"

column_headers = [
    "Year", "Month", "Day", "Hour", "Min", "Sec", "Msec",
    "Bx[nT]", "By[nT]", "Bz[nT]", "Vx[km/s]", "Vy[km/s]", "Vz[km/s]",
    "N[cm^(-3)]", "T[Kelvin]", "MLT", "ML", "[mW m^-2]"
]

# Create an empty excel with these columns
wb_combined = openpyxl.Workbook()
ws_combined = wb_combined.active

# Add the headers to the first row of the worksheet
for col, header in enumerate(column_headers, start=1):
    ws_combined.cell(row=1, column=col, value=header)

#read IMF file
df_IMF = read_IMF_file()

#find all of eflux files in the folder
eflux_files = list_files_in_directory(file_dir_eFlux)

#add elfux data into IMF data
for file in eflux_files:
    #set the eflux file address
    file_path_eFlux = f"{file_dir_eFlux}\{file}" 
    
    #find the date in the filename
    date = extract_datetime_info(file)
    print(  "year = ", date['year'],
            "month = ", date['month'],
            "day = ", date['day'], 
            "hour = ", date["hour"], 
            "minute = ", date['minute'])
    
    #find the IMF data that matches the date
    imf_filtered = df_IMF[(df_IMF['Hour'] == date["hour"]) & 
                          (df_IMF['Min'] == date['minute']) & 
                          (df_IMF['Day'] == date['day']) &
                          (df_IMF['Month'] == date['month']) &
                          (df_IMF['Year'] == date['year'])]
    data_IMF = imf_filtered.values[0]

    #read the eflux file
    df_eflux = read_eflux_file(file_path_eFlux)

    #combine IMF data with eflux data
    for row in df_eflux.itertuples(index=False):
        new_row = []
        new_row = data_IMF
        eflux_row = namedtuples_to_array(row)
        new_row = np.append(new_row, eflux_row)
        ws_combined.append(new_row.tolist())    
        

# Save the workbook to a file
wb_combined.save(fr"{file_dir_main}\combined data.xlsx")
#wb_eflux.save(fr"{file_dir_main}\eflux_test.xlsx")